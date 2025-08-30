from __future__ import annotations
import os, importlib
from typing import Any, List, Dict
import frappe

def _kb_path_for_site() -> str:
    p = frappe.get_site_path("private", "files", "ai_kb")
    os.makedirs(p, exist_ok=True)
    return p

@frappe.whitelist()
def ingest(paths: List[str] | str, collection: str = "frappe_docs") -> Dict[str, Any]:
    if isinstance(paths, str):
        paths = [p.strip() for p in paths.split(",") if p.strip()]
    kb = _kb_path_for_site()
    os.environ["AI_AGENT_KB_PATH"] = kb
    from ai_agent.knowledge import rag_store as rag
    importlib.reload(rag)
    from ai_agent.knowledge import ingest_docs as ing

    store = rag.RAGStore(rag.KBConfig(path=kb, collection=collection))
    files = ing.gather_files(paths)
    batch, n = [], 0
    for f in files:
        suffix = f.suffix.lower()
        chunks = ing._markdown_chunks(f) if suffix in {".md", ".mdx", ".markdown"} else ing._python_doc_chunks(f)
        for text, meta in chunks:
            doc_id = rag._hash_id(f"{meta['source']}::{meta.get('symbol','')}::{meta['chunk']}")
            batch.append((doc_id, text, meta))
        if len(batch) >= 500:
            store.upsert(batch); n += len(batch); batch = []
    if batch:
        store.upsert(batch); n += len(batch)
    return {"kb_path": kb, "files": len(files), "chunks_upserted": n, "collection": collection}

@frappe.whitelist()
def ask(question: str, k: int = 6) -> Dict[str, Any]:
    kb = _kb_path_for_site()
    os.environ["AI_AGENT_KB_PATH"] = kb
    from ai_agent.knowledge import rag_store as rag
    from ai_agent.knowledge import qa as qa_mod
    importlib.reload(rag)
    importlib.reload(qa_mod)
    out = qa_mod.answer_question(question, top_k=int(k))
    return {"answer": out.get("answer"), "sources": out.get("sources", [])}

@frappe.whitelist()
def index_site_schema(collection: str = "frappe_docs") -> Dict[str, Any]:
    kb = _kb_path_for_site()
    os.environ["AI_AGENT_KB_PATH"] = kb
    from ai_agent.knowledge import rag_store as rag
    importlib.reload(rag)
    store = rag.RAGStore(rag.KBConfig(path=kb, collection=collection))
    rows = frappe.get_all("DocType", fields=["name"], limit=5000)
    docs = []
    for r in rows:
        meta = frappe.get_meta(r["name"])
        fields = [f"{df.fieldname} ({df.fieldtype}) - {df.label or ''}" for df in meta.fields]
        text = f"DocType: {meta.name}\nFields: " + ", ".join(fields)
        docs.append((
            rag._hash_id(f"site://doctype::{meta.name}"),
            text,
            {"source": f"site://doctype/{meta.name}", "type": "site_meta", "chunk": 0}
        ))
    store.upsert(docs)
    return {"indexed": len(docs), "kb_path": kb, "collection": collection}

@frappe.whitelist()
def plan_and_execute(command_text: str, dry_run: int = 1, confirm_token: str = ""):
    """
    Plan (and optionally execute) a natural-language command.
    - dry_run=1 → return plan only; no write calls
    - dry_run=0 → execute planned tool calls (will request confirmations for risky ops)
    """
    # Import inside the function to avoid dragging planner on every import
    from .agent import smart_execute as _execute
    user = frappe.session.user if hasattr(frappe, "session") and frappe.session else "Administrator"
    return _execute(command_text, user, dry_run=bool(int(dry_run)), confirm_token=confirm_token)

# ---------- Rich formatting, export & email ----------
import io, csv, json, datetime as dt
from typing import List, Dict, Any, Tuple
from frappe.utils.pdf import get_pdf

def _now_ts() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d_%H%M%S")

def _standardize(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Accepts the agent tool output (e.g., get_purchase_stats/get_sales_stats dicts)
    and returns a consistent structure:
      {
        "title": str,
        "subtitle": str,
        "columns": [str,...],
        "rows": [ [..], ... ],
        "chart": {"labels":[...], "values":[...], "type":"bar"|"line"}
      }
    """
    title = payload.get("title") or "Result"
    subtitle = ""

    # Try common analytics shape: rows=[{label,total}, ...]
    rows_dict = payload.get("rows") or []
    columns = []
    rows = []
    chart = None

    if isinstance(rows_dict, list) and rows_dict and isinstance(rows_dict[0], dict):
        # pick stable columns
        key_order = list(rows_dict[0].keys())
        # promote label & total to the front if present
        for k in ["label", "total"]:
            if k in key_order:
                key_order.remove(k)
        columns = ["label", "total"] + key_order

        for r in rows_dict:
            rows.append([r.get(c, "") for c in columns])

        # simple chart if label+total present
        if "label" in columns and "total" in columns:
            lab_idx, val_idx = columns.index("label"), columns.index("total")
            chart = {
                "type": "bar",
                "labels": [r[lab_idx] for r in rows],
                "values": [float(r[val_idx] or 0) for r in rows],
            }
            # subtitle for “highest …”
            hi = payload.get("highest")
            if hi:
                subtitle = f"Highest: {hi.get('label')} → {hi.get('total')}"
    else:
        # Fallback: stringify payload
        columns = ["key", "value"]
        rows = [[k, json.dumps(v, ensure_ascii=False)[:500]] for k, v in payload.items()]

    # Optional metadata
    meta_bits = []
    if payload.get("from_date") and payload.get("to_date"):
        meta_bits.append(f"{payload['from_date']} → {payload['to_date']}")
    if payload.get("group_by"):
        meta_bits.append(f"grouped by {payload['group_by']}")
    if meta_bits:
        subtitle = (subtitle + (" | " if subtitle else "") + ", ".join(meta_bits)).strip()

    return {
        "title": title,
        "subtitle": subtitle,
        "columns": columns,
        "rows": rows,
        "chart": chart,
        "raw": payload,
    }

def _render_html(std: Dict[str, Any]) -> str:
    cols = std["columns"]
    rows = std["rows"]
    # Basic table; chart rendered on client (JS). This HTML is also used for PDF (no chart there).
    table_html = ["<table style='width:100%;border-collapse:collapse'>",
                  "<thead><tr>"]
    for c in cols:
        table_html.append(f"<th style='border:1px solid #ddd;padding:6px;text-align:left'>{frappe.utils.escape_html(c.title())}</th>")
    table_html.append("</tr></thead><tbody>")
    for r in rows[:1000]:
        table_html.append("<tr>")
        for cell in r:
            table_html.append(f"<td style='border:1px solid #eee;padding:6px'>{frappe.utils.escape_html(str(cell))}</td>")
        table_html.append("</tr>")
    table_html.append("</tbody></table>")

    subtitle = f"<div style='color:#666;margin:6px 0 14px 0'>{frappe.utils.escape_html(std.get('subtitle') or '')}</div>"
    return f"""
      <div>
        <h3 style="margin:0 0 6px 0">{frappe.utils.escape_html(std.get('title') or 'Result')}</h3>
        {subtitle}
        {''.join(table_html)}
      </div>
    """

def _to_csv(std: Dict[str, Any]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(std["columns"])
    for r in std["rows"]:
        w.writerow(r)
    return buf.getvalue().encode("utf-8")

def _to_xlsx(std: Dict[str, Any]) -> bytes:
    try:
        import xlsxwriter
    except Exception:
        # graceful fallback: return CSV bytes if xlsxwriter missing
        return _to_csv(std)
    bio = io.BytesIO()
    wb = xlsxwriter.Workbook(bio, {'in_memory': True})
    ws = wb.add_worksheet("Data")
    head = wb.add_format({'bold': True})
    for ci, c in enumerate(std["columns"]):
        ws.write(0, ci, c, head)
    for ri, row in enumerate(std["rows"], start=1):
        for ci, cell in enumerate(row):
            ws.write(ri, ci, cell)
    wb.close()
    return bio.getvalue()

def _to_pdf(std: Dict[str, Any]) -> bytes:
    html = _render_html(std)
    return get_pdf(html)

def _save_file(fname: str, content: bytes, is_private: bool=True) -> str:
    # creates File doc and returns file_url
    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": fname,
        "is_private": 1 if is_private else 0,
        "content": content,
        "decode": 1,
    }).insert(ignore_permissions=True)
    return file_doc.file_url

@frappe.whitelist()
def run_rich(command_text: str, dry_run: int = 1) -> Dict[str, Any]:
    """
    Plan & (optionally) execute, then return:
      { html, chart, table, raw }
    chart: {labels, values, type}
    table: {columns, rows}
    """
    from .agent import smart_execute as _execute
    user = frappe.session.user if getattr(frappe, "session", None) else "Administrator"
    result = _execute(command_text, user, dry_run=bool(int(dry_run)), confirm_token="")
    std = _standardize(result if isinstance(result, dict) else {"result": result})
    html = _render_html(std)
    table = {"columns": std["columns"], "rows": std["rows"]}
    return {"html": html, "chart": std["chart"], "table": table, "raw": std["raw"]}

@frappe.whitelist()
def export_files(command_text: str, dry_run: int = 0, formats: str = "csv,xlsx,pdf") -> Dict[str, Any]:
    """
    Executes the command, generates requested files, returns file_urls.
    formats: comma separated subset of csv,xlsx,pdf
    """
    from .agent import smart_execute as _execute
    user = frappe.session.user if getattr(frappe, "session", None) else "Administrator"
    result = _execute(command_text, user, dry_run=bool(int(dry_run)), confirm_token="")
    std = _standardize(result if isinstance(result, dict) else {"result": result})
    out = {}
    base = command_text.strip().replace(" ", "_")[:40] or "report"
    ts = _now_ts()

    req = [f.strip().lower() for f in formats.split(",") if f.strip()]
    if "csv" in req:
        url = _save_file(f"{base}_{ts}.csv", _to_csv(std))
        out["csv"] = url
    if "xlsx" in req:
        url = _save_file(f"{base}_{ts}.xlsx", _to_xlsx(std))
        out["xlsx"] = url
    if "pdf" in req:
        url = _save_file(f"{base}_{ts}.pdf", _to_pdf(std))
        out["pdf"] = url
    return {"files": out, "raw": std["raw"]}

@frappe.whitelist()
def run_and_email(command_text: str, to: str, subject: str = "", message: str = "", formats: str = "csv,pdf") -> Dict[str, Any]:
    """
    Executes, exports files, and emails them immediately.
    """
    subj = subject or f"Report: {command_text}"
    files_info = export_files(command_text, dry_run=0, formats=formats)
    files = files_info["files"]

    # Pull file contents for attachments
    atts = []
    for label, url in files.items():
        # Download from File doc directly
        file_doc = frappe.db.get_value("File", {"file_url": url}, ["name", "file_name", "content", "is_private"], as_dict=True)
        # content may not be stored inline; fallback to read from path
        data = None
        if file_doc and file_doc.get("content"):
            data = file_doc["content"]
        else:
            fpath = frappe.utils.get_site_path(url.lstrip("/"))  # /private/files/...
            with open(fpath, "rb") as f:
                data = f.read()
        atts.append({"fname": file_doc["file_name"] if file_doc else url.rsplit("/",1)[-1], "fcontent": data})

    frappe.sendmail(
        recipients=[to],
        subject=subj,
        message=message or f"Attached are the results for: <b>{frappe.utils.escape_html(command_text)}</b>",
        attachments=atts,
    )
    return {"emailed_to": to, "subject": subj, "files": files}

# === Better result extraction + pretty HTML ===
from typing import Iterable

def _extract_best(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Accepts either a raw tool payload (has rows) or an agent wrapper
    (has 'results' list). Returns the best inner result (with rows).
    """
    if isinstance(payload, dict):
        if isinstance(payload.get("rows"), list):
            return payload
        res_list = payload.get("results")
        if isinstance(res_list, Iterable):
            for step in reversed(list(res_list)):
                if isinstance(step, dict):
                    inner = step.get("result")
                    if isinstance(inner, dict) and isinstance(inner.get("rows"), list):
                        return inner
    return payload

def _fmt_money(val) -> str:
    try:
        v = float(val or 0)
    except Exception:
        return str(val)
    try:
        cur = frappe.db.get_single_value("Global Defaults", "default_currency") or None
        from frappe.utils import fmt_money
        return fmt_money(v, currency=cur)
    except Exception:
        return f"{v:,.2f}"

def _standardize(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Take the best inner result, standardize to {title, subtitle, columns, rows, chart, raw}
    """
    data = _extract_best(payload) if isinstance(payload, dict) else payload
    title = data.get("title") or "Result"
    from_date = data.get("from_date")
    to_date = data.get("to_date")
    group_by = (data.get("group_by") or "").strip() or None

    rows = data.get("rows") or []
    # normalize rows of dicts into [label,total,...] with stable column order
    columns: List[str] = []
    table_rows: List[List[Any]] = []

    if rows and isinstance(rows[0], dict):
        # sort by total desc if available
        if "total" in rows[0]:
            try:
                rows = sorted(rows, key=lambda r: float(r.get("total") or 0), reverse=True)
            except Exception:
                pass
        base_cols = list(rows[0].keys())
        # promote label + total to the front if present
        for k in ["label", "total"]:
            if k in base_cols:
                base_cols.remove(k)
        columns = ["label", "total"] + base_cols
        for r in rows:
            table_rows.append([r.get(c, "") for c in columns])
    else:
        # fallback: simple kv table
        columns = ["key", "value"]
        table_rows = [[k, json.dumps(v, ensure_ascii=False)[:800]] for k, v in (data.items() if isinstance(data, dict) else [])]

    # subtitle & quick stats
    meta = []
    if from_date and to_date:
        meta.append(f"{from_date} → {to_date}")
    if group_by:
        meta.append(f"grouped by {group_by.title()}")
    highest = None
    if "label" in columns and "total" in columns and table_rows:
        li, ti = columns.index("label"), columns.index("total")
        try:
            vals = [(r[li], float(r[ti] or 0)) for r in table_rows]
            if vals:
                highest = max(vals, key=lambda t: t[1])
        except Exception:
            highest = None

    subtitle = " · ".join(meta)
    # build chart only if we have label/total
    chart = None
    if "label" in columns and "total" in columns:
        li, ti = columns.index("label"), columns.index("total")
        chart = {
            "type": "bar",
            "labels": [r[li] for r in table_rows][:20],
            "values": [float(r[ti] or 0) for r in table_rows][:20],
        }

    return {
        "title": title,
        "subtitle": subtitle,
        "columns": columns,
        "rows": table_rows,
        "chart": chart,
        "highest": {"label": highest[0], "total": highest[1]} if highest else None,
        "raw": data,
    }

def _render_html(std: Dict[str, Any]) -> str:
    # inline styles to keep it portable (looks good in PDF/email too)
    css = """
    <style>
      .ai-badges{margin:6px 0 12px 0;display:flex;gap:8px;flex-wrap:wrap}
      .ai-badge{background:#f3f4f6;border:1px solid #e5e7eb;border-radius:999px;padding:3px 10px;font-size:12px;color:#374151}
      .ai-stat{display:inline-block;margin:0 12px 12px 0;padding:8px 10px;border-radius:10px;background:#f9fafb;border:1px solid #eee}
      .ai-table{width:100%;border-collapse:collapse;margin-top:8px}
      .ai-table th{background:#fafafa;border:1px solid #eee;padding:6px;text-align:left;font-weight:600}
      .ai-table td{border:1px solid #f0f0f0;padding:6px}
    </style>
    """
    title = frappe.utils.escape_html(std.get("title") or "Result")
    subtitle = frappe.utils.escape_html(std.get("subtitle") or "")
    hi_html = ""
    if std.get("highest"):
        hi = std["highest"]
        hi_html = f"<span class='ai-stat'><b>Highest</b>: {frappe.utils.escape_html(str(hi['label']))} &nbsp;→&nbsp; {_fmt_money(hi['total'])}</span>"

    # pretty table; format "total" column as currency
    cols = std["columns"]
    rows = std["rows"]
    table = ["<table class='ai-table'><thead><tr>"]
    for c in cols:
        table.append(f"<th>{frappe.utils.escape_html(c.title())}</th>")
    table.append("</tr></thead><tbody>")
    for r in rows[:1000]:
        table.append("<tr>")
        for i, cell in enumerate(r):
            if cols[i] == "total":
                cell = _fmt_money(cell)
            table.append(f"<td>{frappe.utils.escape_html(str(cell))}</td>")
        table.append("</tr>")
    table.append("</tbody></table>")

    # badges from subtitle bits
    badges = ""
    if subtitle:
        for bit in subtitle.split(" · "):
            badges += f"<span class='ai-badge'>{frappe.utils.escape_html(bit)}</span>"

    return f"""{css}
      <div>
        <h3 style="margin:0 0 4px 0;font-weight:600">{title}</h3>
        <div class="ai-badges">{badges}</div>
        {hi_html}
        {''.join(table)}
      </div>
    """

@frappe.whitelist()
def run_rich(command_text: str, dry_run: int = 1) -> Dict[str, Any]:
    """
    Plan & (optionally) execute, then return pretty HTML + chart + table.
    """
    from .agent import smart_execute as _execute
    user = frappe.session.user if getattr(frappe, "session", None) else "Administrator"
    raw = _execute(command_text, user, dry_run=bool(int(dry_run)), confirm_token="")
    payload = raw if isinstance(raw, dict) else {"result": raw}
    std = _standardize(payload)
    html = _render_html(std)
    table = {"columns": std["columns"], "rows": std["rows"]}
    return {"html": html, "chart": std["chart"], "table": table, "raw": payload}

# --- AI Agent: exports, email, and presets -----------------------------------
import io, csv, base64, json, datetime
from typing import List, Any
import frappe

def _ensure_list(x):
    if isinstance(x, str):
        try:
            return frappe.parse_json(x)
        except Exception:
            pass
    return x

def _render_table_html(title: str, columns: List[str], rows: List[List[Any]]) -> str:
    # minimal, print-friendly HTML for PDF export
    safe_title = frappe.utils.escape_html(title or "Report")
    head = """
    <style>
      body{font-family:Inter,system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;color:#111827;margin:16px}
      h2{margin:0 0 10px 0}
      table{border-collapse:collapse;width:100%}
      th{background:#f3f4f6;text-align:left;padding:8px;border:1px solid #e5e7eb;font-weight:600}
      td{padding:8px;border:1px solid #f3f4f6}
      .meta{margin:6px 0 14px 0;color:#6b7280;font-size:12px}
    </style>
    """
    hdr = f"<h2>{safe_title}</h2><div class='meta'>Generated: {frappe.utils.now_datetime().strftime('%Y-%m-%d %H:%M:%S')}</div>"
    thead = "<tr>" + "".join(f"<th>{frappe.utils.escape_html(c or '')}</th>" for c in (columns or [])) + "</tr>"
    trows = []
    for r in (rows or []):
        trows.append("<tr>" + "".join(f"<td>{frappe.utils.escape_html('' if v is None else str(v))}</td>" for v in r) + "</tr>")
    return f"<!doctype html><html><head>{head}</head><body>{hdr}<table>{thead}{''.join(trows)}</table></body></html>"

def _bytes_csv(title: str, columns: List[str], rows: List[List[Any]]) -> bytes:
    buf = io.StringIO(newline="")
    w = csv.writer(buf)
    w.writerow(columns or [])
    for r in (rows or []):
        w.writerow(["" if v is None else v for v in r])
    return buf.getvalue().encode("utf-8")

def _bytes_xlsx(title: str, columns: List[str], rows: List[List[Any]]) -> bytes:
    # OpenPyXL is bundled with Frappe
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet1")[:31]
    ws.append(columns or [])
    for r in (rows or []):
        ws.append(["" if v is None else v for v in r])

    # auto width (simple)
    for i, col in enumerate(columns or [], start=1):
        width = max(10, min(50, max([len(str(col))] + [len(str(r[i-1])) if len(r) >= i and r[i-1] is not None else 0 for r in rows or []]) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def _bytes_pdf(title: str, columns: List[str], rows: List[List[Any]]) -> bytes:
    from frappe.utils.pdf import get_pdf
    html = _render_table_html(title, columns, rows)
    return get_pdf(html)

def _save_private_file(filename: str, content: bytes) -> dict:
    """Create a private File and return its info."""
    b64 = base64.b64encode(content).decode()
    f = frappe.get_doc({
        "doctype": "File",
        "file_name": filename,
        "content": b64,
        "decode": True,
        "is_private": 1,
    }).insert(ignore_permissions=True)
    return {"file_url": f.file_url, "file_name": f.file_name, "name": f.name}

@frappe.whitelist()
def list_presets():
    """Small, useful starter prompts for the console."""
    return [
        {"label": "Top 10 customers (YTD)", "prompt": "top 10 customers by sales this year"},
        {"label": "Sales by month (L12M)", "prompt": "sales by month last 12 months"},
        {"label": "Vendor spend (L12M)", "prompt": "purchases by supplier last 12 months"},
        {"label": "Top 10 items (YTD)", "prompt": "top 10 items by sales this year"},
        {"label": "Region split (YTD)", "prompt": "sales by region this year"},
        {"label": "Inventory snapshot", "prompt": "inventory snapshot by warehouse"},
        {"label": "Receivables aging", "prompt": "run report Accounts Receivable Summary for this month"},
        {"label": "Daily run-rate", "prompt": "sales daily run rate vs target this month"},
    ]

@frappe.whitelist()
def export_data(fmt: str, title: str, columns, rows):
    """
    Export the current table to CSV/XLSX/PDF and return File info.
    Args:
      fmt: "csv" | "xlsx" | "pdf"
      title: str
      columns: list[str] (or JSON string)
      rows: list[list] (or JSON string)
    """
    fmt = (fmt or "xlsx").lower().strip()
    columns = _ensure_list(columns) or []
    rows = _ensure_list(rows) or []

    if fmt == "csv":
        content = _bytes_csv(title, columns, rows)
        filename = f"{frappe.scrub(title) or 'report'}.csv"
    elif fmt == "xlsx":
        content = _bytes_xlsx(title, columns, rows)
        filename = f"{frappe.scrub(title) or 'report'}.xlsx"
    elif fmt == "pdf":
        content = _bytes_pdf(title, columns, rows)
        filename = f"{frappe.scrub(title) or 'report'}.pdf"
    else:
        frappe.throw(f"Unsupported format: {fmt}")

    info = _save_private_file(filename, content)
    return {"ok": True, **info}

@frappe.whitelist()
def email_data(to: str, subject: str, message: str, fmt: str, title: str, columns, rows, cc=None, bcc=None):
    """
    Email the current table as an attachment (also saves a private File copy).
    """
    columns = _ensure_list(columns) or []
    rows = _ensure_list(rows) or []

    fmt = (fmt or "xlsx").lower().strip()
    if fmt == "csv":
        content = _bytes_csv(title, columns, rows); ext = "csv"
    elif fmt == "xlsx":
        content = _bytes_xlsx(title, columns, rows); ext = "xlsx"
    elif fmt == "pdf":
        content = _bytes_pdf(title, columns, rows); ext = "pdf"
    else:
        frappe.throw(f"Unsupported format: {fmt}")

    filename = f"{frappe.scrub(title) or 'report'}.{ext}"
    # Save file so recipients can re-download
    info = _save_private_file(filename, content)

    # Send email
    frappe.sendmail(
        recipients=[to] if isinstance(to, str) else to,
        cc=cc or None, bcc=bcc or None,
        subject=subject or title or "Report",
        message=message or f"Attached is {title or 'your report'}.\n\nDownload: {info['file_url']}",
        attachments=[{"fname": filename, "fcontent": content}],
    )
    return {"ok": True, "file_url": info["file_url"], "file_name": info["file_name"]}
